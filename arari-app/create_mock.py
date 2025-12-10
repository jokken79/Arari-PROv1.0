
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
# Create sheet with "wrong" name to test intelligent search
ws = wb.active
ws.title = "Sheet1"

# Add some noise rows
ws.cell(row=1, column=1, value="Company Confidential")
ws.cell(row=2, column=1, value="Generated on 2025-01-01")

# Add Header at Row 5
headers = ["No", "Name", "Kana", "Status", "Company", "Billing"]
ws.cell(row=5, column=1, value="No")
ws.cell(row=5, column=2, value="氏名")
ws.cell(row=5, column=3, value="フリガナ")
ws.cell(row=5, column=4, value="ステータス")
ws.cell(row=5, column=5, value="派遣先")
ws.cell(row=5, column=6, value="請求単価")

# Add Data
data = [
    (101, "Yamada Taro", "ヤマダ タロウ", "在籍", "Toyota", 2500),
    (102, "Suzuki Hanako", "スズキ ハナコ", "在籍", "Honda", 3000),
    (103, "Tanaka Ken", "タナカ ケン", "退職", "Sony", 2800),
]

for i, row in enumerate(data):
    for j, val in enumerate(row):
        ws.cell(row=6+i, column=j+1, value=val)

wb.save("mock_employees.xlsx")
print("Created mock_employees.xlsx")
