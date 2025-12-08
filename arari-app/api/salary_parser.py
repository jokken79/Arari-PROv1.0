"""
Specialized parser for .xlsm salary statement files (給与明細)

Handles complex multi-sheet, multi-employee layout where:
- Each file has multiple sheets (1 summary + 25 company sheets)
- Each company sheet has multiple employees side-by-side
- Each employee occupies ~14 columns with data in fixed row positions
"""

import openpyxl
import re
from typing import List, Optional
from io import BytesIO
from models import PayrollRecordCreate
from employee_rates import get_rates_loader


class SalaryStatementParser:
    """Parser for complex .xlsm salary statement files (給与明細)"""

    # Each employee occupies this many columns width
    EMPLOYEE_COLUMN_WIDTH = 14

    # Row positions for data extraction (1-indexed)
    # Verified from actual file structure
    ROW_POSITIONS = {
        'period': 5,          # "2025年1月分(2月17日支給)"
        'employee_id': 6,     # 6-digit employee ID
        'name': 7,            # Employee name
        'work_days': 11,      # 出勤日数
        'paid_leave_days': 12,  # 有給日数 (row 12, offset 5)
        'work_hours': 13,     # 労働時間
        'overtime_hours': 14, # 残業時間
        'base_salary': 16,    # 基本給
        'overtime_pay': 17,   # 残業代
        'other_allowances': 18,  # その他手当
        'paid_leave_payment': 21,  # 有給休暇支給額
        'transport_allowance': 22,  # 通勤費
        'gross_salary': 30,   # 支給合計
        'health_insurance': 31,    # 健康保険料
        'pension': 32,             # 厚生年金
        'employment_insurance': 33, # 雇用保険料
        'social_insurance': 34,    # 社会保険料計 (TOTAL)
        'resident_tax': 35,        # 住民税
        'net_salary': 36,          # 差引支給額 (final net)
    }

    # Column offsets within an employee block
    # First employee block starts at col 1, employee_id at col 10
    # So offset = 9
    # Period is at col 3 in this example = offset 2
    # Name is at col 3 = offset 2
    # New fields (rows 31-36) use offset 3 (same as other salary data)
    COLUMN_OFFSETS = {
        'period': 2,          # [3] = col 3 - base 1
        'employee_id': 9,     # [10] = col 10 - base 1
        'name': 2,            # [3] = col 3 - base 1
        'work_days': 5,       # [6] from data = col 6 - base 1
        'paid_leave_days': 5,  # offset 5 (same as work_days offset)
        'work_hours': 3,      # [4] from data = col 4 - base 1
        'overtime_hours': 3,  # [4] from data
        'base_salary': 3,     # [4] from data
        'overtime_pay': 3,    # [4] from data
        'other_allowances': 3,  # [4] from data
        'paid_leave_payment': 3,  # 有給休暇支給額
        'transport_allowance': 3,  # [4] from data
        'gross_salary': 3,    # 支給合計
        'health_insurance': 3,     # 健康保険料
        'pension': 3,              # 厚生年金
        'employment_insurance': 3, # 雇用保険料
        'social_insurance': 3,     # 社会保険料計
        'resident_tax': 3,         # 住民税
        'net_salary': 3,      # 差引支給額
    }

    def parse(self, content: bytes) -> List[PayrollRecordCreate]:
        """
        Parse .xlsm file and extract all employee payroll records

        Args:
            content: Binary content of the Excel file

        Returns:
            List of PayrollRecordCreate objects
        """
        try:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return []

        records = []

        # Process all sheets except summary (集計) and subcontracting (請負)
        # 請負 employees don't generate billing from factories
        for sheet_name in wb.sheetnames:
            if sheet_name == '集計' or sheet_name == 'Summary' or '請負' in sheet_name:
                continue  # Skip summary and subcontracting sheets

            try:
                ws = wb[sheet_name]
                sheet_records = self._parse_sheet(ws, sheet_name)
                records.extend(sheet_records)
            except Exception as e:
                print(f"Warning: Error parsing sheet '{sheet_name}': {e}")
                continue

        return records

    def _parse_sheet(self, ws, sheet_name: str) -> List[PayrollRecordCreate]:
        """
        Parse a single company sheet containing multiple employees side-by-side

        Args:
            ws: openpyxl worksheet
            sheet_name: Name of the sheet (派遣先企業 = dispatch company)

        Returns:
            List of PayrollRecordCreate objects from this sheet
        """
        records = []

        # Detect employee column positions by finding employee IDs in row 6
        employee_cols = self._detect_employee_columns(ws)

        if not employee_cols:
            print(f"  Warning: No employee IDs found in sheet '{sheet_name}'")
            return records

        # Extract data for each employee, passing the sheet_name as dispatch_company
        for col_idx in employee_cols:
            record = self._extract_employee_data(ws, col_idx, sheet_name)
            if record:
                records.append(record)

        return records

    def _detect_employee_columns(self, ws) -> List[int]:
        """
        Find column indices where employee blocks start

        Employee IDs are 6-digit numbers appearing in row 6 at columns: 10, 24, 38, 52, ...
        Each employee block occupies ~14 columns starting from column: 1, 15, 29, 43, ...

        We find employee ID positions, then calculate the base column of each block

        Returns:
            List of base column indices (1-indexed) for each employee block
        """
        columns = []

        # Scan row 6 (employee_id row) for 6-digit numbers
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=self.ROW_POSITIONS['employee_id'], column=col).value

            if cell_value is None:
                continue

            # Check if it's a 6-digit number (employee ID)
            try:
                emp_id_str = str(cell_value).strip()
                if emp_id_str.isdigit() and len(emp_id_str) == 6:
                    # Found an employee ID at column col
                    # The employee IDs appear at columns: 10, 24, 38, 52...
                    # Each block starts 9 columns before the employee ID
                    # So: block 1 starts at 10-9=1, block 2 starts at 24-9=15, etc.
                    base_col = col - 9
                    if base_col > 0 and base_col not in columns:
                        columns.append(base_col)
            except (ValueError, AttributeError):
                continue

        return sorted(columns)

    def _extract_employee_data(self, ws, base_col: int, dispatch_company: str = '') -> Optional[PayrollRecordCreate]:
        """
        Extract data for one employee from a specific column position

        Args:
            ws: openpyxl worksheet
            base_col: Starting column (1-indexed) of the employee block
            dispatch_company: Name of the Excel sheet (派遣先企業)

        Returns:
            PayrollRecordCreate object or None if extraction fails
        """
        try:
            # Extract period from row 5
            period_cell = ws.cell(
                row=self.ROW_POSITIONS['period'],
                column=base_col + self.COLUMN_OFFSETS['period']
            )
            period = self._parse_period(period_cell.value)
            if not period:
                return None

            # Extract employee_id from row 6
            emp_id_cell = ws.cell(
                row=self.ROW_POSITIONS['employee_id'],
                column=base_col + self.COLUMN_OFFSETS['employee_id']
            )
            employee_id = str(emp_id_cell.value or '').strip()

            if not employee_id or not employee_id.isdigit():
                return None

            # Extract employee name from row 7
            name_cell = ws.cell(
                row=self.ROW_POSITIONS['name'],
                column=base_col + self.COLUMN_OFFSETS['name']
            )
            employee_name = str(name_cell.value or '').strip()

            # Load employee rates (時給 and 単価) from 社員台帳
            rates_loader = get_rates_loader()
            hourly_rate, billing_rate = rates_loader.get_rates(employee_id)

            # Extract numeric data from fixed row positions
            paid_leave_days = self._get_numeric(ws, self.ROW_POSITIONS['paid_leave_days'], base_col + self.COLUMN_OFFSETS['paid_leave_days'])

            # Extract deduction fields from Excel (rows 31-36)
            health_insurance = self._get_numeric(ws, self.ROW_POSITIONS['health_insurance'], base_col + self.COLUMN_OFFSETS['health_insurance'])
            pension = self._get_numeric(ws, self.ROW_POSITIONS['pension'], base_col + self.COLUMN_OFFSETS['pension'])
            employment_insurance = self._get_numeric(ws, self.ROW_POSITIONS['employment_insurance'], base_col + self.COLUMN_OFFSETS['employment_insurance'])
            social_insurance_total = self._get_numeric(ws, self.ROW_POSITIONS['social_insurance'], base_col + self.COLUMN_OFFSETS['social_insurance'])
            resident_tax = self._get_numeric(ws, self.ROW_POSITIONS['resident_tax'], base_col + self.COLUMN_OFFSETS['resident_tax'])

            # If social_insurance_total is 0, calculate from components
            if social_insurance_total == 0:
                social_insurance_total = health_insurance + pension + employment_insurance

            # Calculate total hours for billing
            total_hours = self._get_numeric(ws, self.ROW_POSITIONS['work_hours'], base_col + self.COLUMN_OFFSETS['work_hours'])
            total_hours += self._get_numeric(ws, self.ROW_POSITIONS['overtime_hours'], base_col + self.COLUMN_OFFSETS['overtime_hours'])

            data = {
                'employee_id': employee_id,
                'period': period,
                'dispatch_company': dispatch_company,  # 派遣先企業 (Excel sheet name)
                'employee_name': employee_name,  # 氏名 from Excel
                'work_days': self._get_numeric(ws, self.ROW_POSITIONS['work_days'], base_col + self.COLUMN_OFFSETS['work_days']),
                'paid_leave_days': paid_leave_days,
                'work_hours': self._get_numeric(ws, self.ROW_POSITIONS['work_hours'], base_col + self.COLUMN_OFFSETS['work_hours']),
                'overtime_hours': self._get_numeric(ws, self.ROW_POSITIONS['overtime_hours'], base_col + self.COLUMN_OFFSETS['overtime_hours']),
                'base_salary': self._get_numeric(ws, self.ROW_POSITIONS['base_salary'], base_col + self.COLUMN_OFFSETS['base_salary']),
                'overtime_pay': self._get_numeric(ws, self.ROW_POSITIONS['overtime_pay'], base_col + self.COLUMN_OFFSETS['overtime_pay']),
                'other_allowances': self._get_numeric(ws, self.ROW_POSITIONS['other_allowances'], base_col + self.COLUMN_OFFSETS['other_allowances']),
                'transport_allowance': self._get_numeric(ws, self.ROW_POSITIONS['transport_allowance'], base_col + self.COLUMN_OFFSETS['transport_allowance']),
                'net_salary': self._get_numeric(ws, self.ROW_POSITIONS['net_salary'], base_col + self.COLUMN_OFFSETS['net_salary']),

                # Deduction fields extracted from Excel
                'paid_leave_hours': paid_leave_days * 8,  # Calculate hours from days
                'social_insurance': social_insurance_total,  # 社会保険料計 from Excel
                'employment_insurance': employment_insurance,  # 雇用保険料 from Excel
                'income_tax': 0,  # Not typically in this Excel format
                'resident_tax': resident_tax,  # 住民税 from Excel
                'other_deductions': 0,
                # Calculate billing amount from employee's billing_rate (単価) and total hours
                'billing_amount': billing_rate * total_hours if billing_rate > 0 else 0,
            }

            # Read gross_salary directly from Excel (row 30)
            gross_salary_from_excel = self._get_numeric(ws, self.ROW_POSITIONS['gross_salary'], base_col + self.COLUMN_OFFSETS['gross_salary'])

            # Use Excel value if available, otherwise calculate
            if gross_salary_from_excel > 0:
                data['gross_salary'] = gross_salary_from_excel
            else:
                data['gross_salary'] = (
                    data['base_salary'] +
                    data['overtime_pay'] +
                    data['other_allowances'] +
                    data['transport_allowance']
                )

            # Create and return PayrollRecordCreate
            return PayrollRecordCreate(**data)

        except Exception as e:
            # Silently return None - employee data may not be valid at this position
            return None

    def _parse_period(self, value) -> str:
        """
        Convert period string format '2025年1月分(2月17日支給)' to '2025年1月'

        Args:
            value: Raw period value from Excel cell

        Returns:
            Formatted period string (YYYY年M月) or empty string if parsing fails
        """
        if value is None or value == '':
            return ''

        value_str = str(value)

        # Match pattern: 年 followed by digits, then 月
        match = re.search(r'(\d{4})年(\d{1,2})月', value_str)
        if match:
            year = match.group(1)
            month = match.group(2).zfill(2)  # Pad with zero if needed
            return f"{year}年{int(month)}月"  # Remove padding for standard format

        return ''

    def _get_numeric(self, ws, row: int, col: int) -> float:
        """
        Safely extract numeric value from a cell

        Args:
            ws: openpyxl worksheet
            row: Row number (1-indexed)
            col: Column number (1-indexed)

        Returns:
            Float value or 0.0 if cell is empty or not numeric
        """
        try:
            cell = ws.cell(row=row, column=col)
            value = cell.value

            if value is None or value == '':
                return 0.0

            # Handle different numeric types
            if isinstance(value, (int, float)):
                return float(value)

            # Try to convert string to float
            value_str = str(value).strip()
            if not value_str:
                return 0.0

            return float(value_str)

        except (ValueError, TypeError, AttributeError):
            # Return 0 if conversion fails
            return 0.0
